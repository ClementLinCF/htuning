import openpyxl
from openpyxl.styles import PatternFill
import pandas as pd

wb = openpyxl.load_workbook('comparison.xlsx')

improve_ws = wb['improve']

for row in improve_ws.iter_rows(min_row=2, max_col=8):
    for cell in row:
        try:
            print(cell.value)
            value = float(cell.value)
            if value < 0:
                cell.value = 0
        except (ValueError, TypeError):
            continue

wb.save('comparison.xlsx')


wb = openpyxl.load_workbook('comparison.xlsx')

final_improve_ws = wb['improve']


untuned_ws = wb['untuned']
untuned_h_col = [row[7].value for row in untuned_ws.iter_rows(min_row=2)]

if final_improve_ws.cell(1, 9).value is None:
    final_improve_ws.cell(1, 9).value = "untuned us"

for i, value in enumerate(untuned_h_col, start=2):
    final_improve_ws.cell(i, 9, value)


freq_result = pd.read_csv('freq_result.csv', header=None)


final_data = [
    [str(cell.value).strip() if cell.value is not None else "" for cell in row[:5]]
    for row in final_improve_ws.iter_rows(min_row=2)
]


print(final_data)
print(len(final_data))
freq_result[[0, 1, 2, 3, 4]] = freq_result[[0, 1, 2, 3, 4]].astype(str)

if final_improve_ws.cell(1, 10).value is None:
    final_improve_ws.cell(1, 10).value = "count"

for i, row in enumerate(final_data, start=2):
    match_row = freq_result[
        (freq_result[0] == str(row[0])) &
        (freq_result[1] == str(row[1])) &
        (freq_result[2] == str(row[2])) &
        (freq_result[3] == str(row[3])) &
        (freq_result[4] == str(row[4]))
    ]
    count_value = match_row[5].values[0] if not match_row.empty else 0
    final_improve_ws.cell(i, 10, count_value)

if final_improve_ws.cell(1, 11).value is None:
    final_improve_ws.cell(1, 11).value = "total latency"


for i in range(2, final_improve_ws.max_row + 1):
    count_value = final_improve_ws.cell(i, 10).value or 0
    untuned_value = final_improve_ws.cell(i, 9).value or 0
    count_value = float(count_value) if isinstance(count_value, (int, float, str)) and str(count_value).replace('.', '', 1).isdigit() else 0
    untuned_value = float(untuned_value) if isinstance(untuned_value, (int, float, str)) and str(untuned_value).replace('.', '', 1).isdigit() else 0

    final_improve_ws.cell(i, 11, count_value * untuned_value)


if final_improve_ws.cell(1, 12).value is None:
    final_improve_ws.cell(1, 12).value = "reduced total latency"

for i in range(2, final_improve_ws.max_row + 1):
    total_latency = final_improve_ws.cell(i, 11).value or 0
    improvement_rate = final_improve_ws.cell(i, 8).value or 0
    print("total_latyenc = ", total_latency)
    print("improvement_rate = ", improvement_rate)
    reduced_latency = total_latency - (total_latency * improvement_rate / 100)
    final_improve_ws.cell(i, 12, reduced_latency)

total_latency_sum = sum(
    final_improve_ws.cell(i, 11).value or 0 for i in range(2, final_improve_ws.max_row + 1)
)
reduced_latency_sum = sum(
    final_improve_ws.cell(i, 12).value or 0 for i in range(2, final_improve_ws.max_row + 1)
)
reduction_percentage = ((total_latency_sum - reduced_latency_sum) / total_latency_sum) * 100 if total_latency_sum else 0

final_row = final_improve_ws.max_row + 1
final_improve_ws.cell(final_row, 12, f"{reduction_percentage:.2f}%")

yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
final_improve_ws.cell(final_row, 12).fill = yellow_fill

wb.save('comparison_updated.xlsx')
print("Processed and saved as 'comparison_updated.xlsx'")

